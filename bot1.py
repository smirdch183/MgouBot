import pandas
import json
import math
import config
import time 
import schedule 
import telebot
from telebot import types
from telebot import datetime
from datetime import timedelta
from openpyxl import load_workbook

token = config.token
bot = telebot.TeleBot(token)

def upgrade():
    wb = load_workbook(filename='45_03_02_Иностранные_языки_и_культуры_стран_изучаемых_языков_английский.xlsx')
    sheet_ranges = wb['Лист1']
    bokv = sheet_ranges.merged_cells.ranges
    from openpyxl.utils.cell import range_boundaries
    for st_name in wb.sheetnames:
        st = wb[st_name]
        mcr_coord_list = [mcr.coord for mcr in st.merged_cells.ranges]
        
        for mcr in mcr_coord_list:
            min_col, min_row, max_col, max_row = range_boundaries(mcr)
            top_left_cell_value = st.cell(row=min_row, column=min_col).value
            st.unmerge_cells(mcr)
            for row in st.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
                for cell in row:
                    cell.value = top_left_cell_value

    wb.save('Gotov.xlsx')

@bot.message_handler(commands=['start'])
def start_message(message):
    
    with open('groups.json', 'r') as f:
        groupload = json.load(f)
    
    if str(message.chat.id) not in groupload:
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
        btn1 = types.KeyboardButton('04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).1')
        btn2 = types.KeyboardButton('04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).2')
        btn3 = types.KeyboardButton('04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).4')
        btn4 = types.KeyboardButton('04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).5')
        btn5 = types.KeyboardButton('04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).6')
        btn6 = types.KeyboardButton('04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).1')
        btn7 = types.KeyboardButton('04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).2')
        btn8 = types.KeyboardButton('04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).4')
        btn9 = types.KeyboardButton('04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).5')
        btn10 = types.KeyboardButton('04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).1')
        btn11 = types.KeyboardButton('04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).2')
        btn12 = types.KeyboardButton('04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).3')
        btn13 = types.KeyboardButton('04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).4')
        btn14 = types.KeyboardButton('04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).5')
        markup.add(btn1,btn2,btn3,btn4,btn5,btn6,btn7,btn8,btn9,btn10,btn11,btn12,btn13,btn14)
        bot.send_message(message.chat.id,"Привет, я МГОПУ бот🤖\nВыбери свою группу", reply_markup=markup)
    else:
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
        btn1 = types.KeyboardButton('04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).1')
        btn2 = types.KeyboardButton('04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).2')
        btn3 = types.KeyboardButton('04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).4')
        btn4 = types.KeyboardButton('04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).5')
        btn5 = types.KeyboardButton('04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).6')
        btn6 = types.KeyboardButton('04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).1')
        btn7 = types.KeyboardButton('04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).2')
        btn8 = types.KeyboardButton('04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).4')
        btn9 = types.KeyboardButton('04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).5')
        btn10 = types.KeyboardButton('04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).1')
        btn11 = types.KeyboardButton('04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).2')
        btn12 = types.KeyboardButton('04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).3')
        btn13 = types.KeyboardButton('04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).4')
        btn14 = types.KeyboardButton('04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).5')
        btnmenu = types.KeyboardButton('Меню')
        markup.add(btn1,btn2,btn3,btn4,btn5,btn6,btn7,btn8,btn9,btn10,btn11,btn12,btn13,btn14, btnmenu)
        bot.send_message(message.chat.id,"Привет, я МГОПУ бот🤖\nВыбери свою группу", reply_markup=markup)

@bot.message_handler(content_types=['text'])
def get_text_message(message):
    with open('groups.json', 'r') as f:
        groupload = json.load(f)
    if str(message.chat.id) not in groupload:
        if message.text == '04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).1':
            user_groupid = 1
        elif message.text == '04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).2':
            user_groupid = 2
        elif message.text == '04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).4':
            user_groupid = 3
        elif message.text == '04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).5':
            user_groupid = 4
        elif message.text == '04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).6':
            user_groupid = 5
        elif message.text == '04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).1':
            user_groupid = 6
        elif message.text == '04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).2':
            user_groupid = 7
        elif message.text == '04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).4':
            user_groupid = 8
        elif message.text == '04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).5':
            user_groupid = 9
        elif message.text == '04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).1':
            user_groupid = 10
        elif message.text == '04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).2':
            user_groupid = 11
        elif message.text == '04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).3':
            user_groupid = 12
        elif message.text == '04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).4':
            user_groupid = 13
        elif message.text == '04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).5':
            user_groupid = 14
        else:
            user_groupid = 0
            bot.send_message(message.chat.id,"Такой группы нет, выберите существующую")
        if user_groupid != 0:
            with open('groups.json', 'w') as f:
                user_id = message.chat.id
                groupload[user_id] = {'groupid': user_groupid}
                json.dump(groupload, f, indent=4, ensure_ascii=False)
                markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
                btnScheduleForTomorrow = types.KeyboardButton('Расписание на завтра')
                btnSheduleForTodayy = types.KeyboardButton('Расписание на сегодня')
                btnSheduleForDate = types.KeyboardButton('Расписание на дату')
                btnUpdateTable = types.KeyboardButton('Обновить таблицу')
                btnEditGroup = types.KeyboardButton('Изменить группу')
                btnError = types.KeyboardButton('Сообщить об ошибке')
                btnDonat = types.KeyboardButton('Пожертвование')
                markup.add(btnScheduleForTomorrow, btnSheduleForTodayy, btnSheduleForDate, btnUpdateTable, btnEditGroup, btnError, btnDonat)
                bot.send_message(message.chat.id,"Данные сохранены", reply_markup=markup)
        else:
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
            btn1 = types.KeyboardButton('04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).1')
            btn2 = types.KeyboardButton('04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).2')
            btn3 = types.KeyboardButton('04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).4')
            btn4 = types.KeyboardButton('04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).5')
            btn5 = types.KeyboardButton('04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).6')
            btn6 = types.KeyboardButton('04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).1')
            btn7 = types.KeyboardButton('04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).2')
            btn8 = types.KeyboardButton('04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).4')
            btn9 = types.KeyboardButton('04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).5')
            btn10 = types.KeyboardButton('04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).1')
            btn11 = types.KeyboardButton('04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).2')
            btn12 = types.KeyboardButton('04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).3')
            btn13 = types.KeyboardButton('04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).4')
            btn14 = types.KeyboardButton('04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).5')
            markup.add(btn1,btn2,btn3,btn4,btn5,btn6,btn7,btn8,btn9,btn10,btn11,btn12,btn13,btn14)
            bot.send_message(message.chat.id,"Выберети группу", reply_markup=markup)
        with open('groups.json', 'r') as f:
            groupload = json.load(f)
    else:
        error = [f'ошибка']
        for errors in error:
            if errors in message.text.lower(): 
                bot.send_message(615009766,'{0}: {1}'.format(message.chat.username,message.text))
                bot.send_message(message.chat.id,'Сообщение отправлено')
        if message.text == 'Обновить таблицу':
            bot.send_message(message.chat.id,"Обновление...")
            upgrade()
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            btnScheduleForTomorrow = types.KeyboardButton('Расписание на завтра')
            btnSheduleForTodayy = types.KeyboardButton('Расписание на сегодня')
            btnSheduleForDate = types.KeyboardButton('Расписание на дату')
            btnUpdateTable = types.KeyboardButton('Обновить таблицу')
            btnEditGroup = types.KeyboardButton('Изменить группу')
            btnError = types.KeyboardButton('Сообщить об ошибке')
            btnDonat = types.KeyboardButton('Пожертвование')
            markup.add(btnScheduleForTomorrow, btnSheduleForTodayy, btnSheduleForDate, btnUpdateTable, btnEditGroup, btnError, btnDonat)
            bot.send_message(message.chat.id,"Таблица обновлена", reply_markup=markup)
        if message.text == 'Пожертвование':
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            btnScheduleForTomorrow = types.KeyboardButton('Расписание на завтра')
            btnSheduleForTodayy = types.KeyboardButton('Расписание на сегодня')
            btnSheduleForDate = types.KeyboardButton('Расписание на дату')
            btnUpdateTable = types.KeyboardButton('Обновить таблицу')
            btnEditGroup = types.KeyboardButton('Изменить группу')
            btnError = types.KeyboardButton('Сообщить об ошибке')
            btnDonat = types.KeyboardButton('Пожертвование')
            markup.add(btnScheduleForTomorrow, btnSheduleForTodayy, btnSheduleForDate, btnUpdateTable, btnEditGroup, btnError, btnDonat)
            bot.send_message(message.chat.id,"Сбер: 5469400038728155\nТинькофф: 2200700158154475", reply_markup=markup)
        if message.text == 'Изменить группу':
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
            btn1 = types.KeyboardButton('04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).1')
            btn2 = types.KeyboardButton('04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).2')
            btn3 = types.KeyboardButton('04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).4')
            btn4 = types.KeyboardButton('04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).5')
            btn5 = types.KeyboardButton('04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).6')
            btn6 = types.KeyboardButton('04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).1')
            btn7 = types.KeyboardButton('04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).2')
            btn8 = types.KeyboardButton('04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).4')
            btn9 = types.KeyboardButton('04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).5')
            btn10 = types.KeyboardButton('04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).1')
            btn11 = types.KeyboardButton('04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).2')
            btn12 = types.KeyboardButton('04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).3')
            btn13 = types.KeyboardButton('04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).4')
            btn14 = types.KeyboardButton('04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).5')
            btnmenu = types.KeyboardButton('Назад')
            markup.add(btn1,btn2,btn3,btn4,btn5,btn6,btn7,btn8,btn9,btn10,btn11,btn12,btn13,btn14, btnmenu)
            bot.send_message(message.chat.id,"Выберите группу", reply_markup=markup)
        date = message.text
        try:
            datetime.strptime(date, '%d.%m.%Y')
            date = True
        except ValueError:
            date = False
        if date == True:
            date = message.text
            data = datetime.strptime(date, '%d.%m.%Y')
            data = data.strftime('%Y-%m-%d')
            excel_data_df = pandas.read_excel(
            'Gotov.xlsx',
            sheet_name='Лист1', skiprows=[0])
            with open('groups.json', 'r') as f:
                groupload = json.load(f)
            chatid = groupload['{0}'.format(message.chat.id)]["groupid"]
            chatid = chatid+2
            if chatid == 3:
                chatBuckv = "04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).1"
            if chatid == 4:
                chatBuckv = "04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).2"
            if chatid == 5:
                chatBuckv = "04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).4"
            if chatid == 6:
                chatBuckv = "04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).5"
            if chatid == 7:
                chatBuckv = "04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).6"
            if chatid == 8:
                chatBuckv = "04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).1"
            if chatid == 9:
                chatBuckv = "04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).2"
            if chatid == 10:
                chatBuckv = "04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).4"
            if chatid == 11:
                chatBuckv = "04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).5"
            if chatid == 12:
                chatBuckv = "04.ЛОБ.21.ИЯИКСИЯ(АЯИИЯ).1"
            if chatid == 13:
                chatBuckv = "04.ЛОБ.21.ИЯИКСИЯ(АЯИИЯ).2"
            if chatid == 14:
                chatBuckv = "04.ЛОБ.21.ИЯИКСИЯ(АЯИИЯ).3"
            if chatid == 15:
                chatBuckv = "04.ЛОБ.21.ИЯИКСИЯ(АЯИИЯ).4"
            if chatid == 16:
                chatBuckv = "04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).5"
            excel_data_df['День'] = excel_data_df['День'].fillna(method='ffill')
            massiv_index = excel_data_df.index [excel_data_df['День']== data ]. tolist()
            if massiv_index == []:
                markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
                btnScheduleForTomorrow = types.KeyboardButton('Расписание на завтра')
                btnSheduleForTodayy = types.KeyboardButton('Расписание на сегодня')
                btnSheduleForDate = types.KeyboardButton('Расписание на дату')
                btnUpdateTable = types.KeyboardButton('Обновить таблицу')
                btnEditGroup = types.KeyboardButton('Изменить группу')
                btnError = types.KeyboardButton('Сообщить об ошибке')
                btnDonat = types.KeyboardButton('Пожертвование')
                markup.add(btnScheduleForTomorrow, btnSheduleForTodayy, btnSheduleForDate, btnUpdateTable, btnEditGroup, btnError, btnDonat)
                bot.send_message(message.chat.id,"Не могу найти эту дату", reply_markup=markup)
            else:
                d1,d2,d3,d4,d5,d6,d7,d8 = excel_data_df['Пара '].loc[excel_data_df.index[massiv_index]]
                i1,i2,i3,i4,i5,i6,i7,i8 = excel_data_df[chatBuckv].loc[excel_data_df.index[massiv_index]]
                raz = excel_data_df[chatBuckv].loc[excel_data_df.index[massiv_index]]
                if type(i1) not in [str]:
                    i1n = math.isnan(i1)
                elif type(i1) == str:
                    i1n = False
                if type(i2) not in [str]:
                    i2n = math.isnan(i2)
                elif type(i2) == str:
                    i2n = False
                if type(i3) not in [str]:
                    i3n = math.isnan(i3)
                elif type(i3) == str:
                    i3n = False
                if type(i4) not in [str]:
                    i4n = math.isnan(i4)
                elif type(i4) == str:
                    i4n = False
                if type(i5) not in [str]:
                    i5n = math.isnan(i5)
                elif type(i5) == str:
                    i5n = False
                if type(i6) not in [str]:
                    i6n = math.isnan(i6)
                elif type(i6) == str:
                    i6n = False
                if type(i7) not in [str]:
                    i7n = math.isnan(i7)
                elif type(i7) == str:
                    i7n = False
                if type(i8) not in [str]:
                    i8n = math.isnan(i8)
                elif type(i8) == str:
                    i8n = False
                if i1n == True & i2n == True & i3n == True & i4n == True & i5n == True & i6n == True & i7n == True & i8n == True:
                    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
                    btnScheduleForTomorrow = types.KeyboardButton('Расписание на завтра')
                    btnSheduleForTodayy = types.KeyboardButton('Расписание на сегодня')
                    btnSheduleForDate = types.KeyboardButton('Расписание на дату')
                    btnUpdateTable = types.KeyboardButton('Обновить таблицу')
                    btnEditGroup = types.KeyboardButton('Изменить группу')
                    btnError = types.KeyboardButton('Сообщить об ошибке')
                    btnDonat = types.KeyboardButton('Пожертвование')
                    markup.add(btnScheduleForTomorrow, btnSheduleForTodayy, btnSheduleForDate, btnUpdateTable, btnEditGroup, btnError, btnDonat)
                    bot.send_message(message.chat.id,"Выходной 🥳", reply_markup=markup)
                else:
                    if type(i1) not in [str]:
                        i1 = 'Нет пары'
                    if type(i2) not in [str]:
                        i2 = 'Нет пары'
                    if type(i3) not in [str]:
                        i3 = 'Нет пары'
                    if type(i4) not in [str]:
                        i4 = 'Нет пары'
                    if type(i5) not in [str]:
                        i5 = 'Нет пары'
                    if type(i6) not in [str]:
                        i6 = 'Нет пары'
                    if type(i7) not in [str]:
                        i7 = 'Нет пары'
                    if type(i8) not in [str]:
                        i8 = 'Нет пары'
                    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
                    btnScheduleForTomorrow = types.KeyboardButton('Расписание на завтра')
                    btnSheduleForTodayy = types.KeyboardButton('Расписание на сегодня')
                    btnSheduleForDate = types.KeyboardButton('Расписание на дату')
                    btnUpdateTable = types.KeyboardButton('Обновить таблицу')
                    btnEditGroup = types.KeyboardButton('Изменить группу')
                    btnError = types.KeyboardButton('Сообщить об ошибке')
                    btnDonat = types.KeyboardButton('Пожертвование')
                    markup.add(btnScheduleForTomorrow, btnSheduleForTodayy, btnSheduleForDate, btnUpdateTable, btnEditGroup, btnError, btnDonat)
                    bot.send_message(message.chat.id,'{0}\n{1}\n\n{2}\n{3}\n\n{4}\n{5}\n\n{6}\n{7}\n\n{8}\n{9}\n\n{10}\n{11}\n\n{12}\n{13}\n\n{14}\n{15}'
                    .format(d1,i1,d2,i2,d3,i3,d4,i4,d5,i5,d6,i6,d7,i7,d8,i8), reply_markup=markup)
        if message.text == 'Расписание на дату':
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            btnScheduleForTomorrow = types.KeyboardButton('Расписание на завтра')
            btnSheduleForTodayy = types.KeyboardButton('Расписание на сегодня')
            btnSheduleForDate = types.KeyboardButton('Расписание на дату')
            btnUpdateTable = types.KeyboardButton('Обновить таблицу')
            btnEditGroup = types.KeyboardButton('Изменить группу')
            btnError = types.KeyboardButton('Сообщить об ошибке')
            btnDonat = types.KeyboardButton('Пожертвование')
            markup.add(btnScheduleForTomorrow, btnSheduleForTodayy, btnSheduleForDate, btnUpdateTable, btnEditGroup, btnError, btnDonat)
            bot.send_message(message.chat.id,"Просто напиши дату на которое нужно узнать расписание.\nПример:\n26.12.2022", reply_markup=markup)
        if message.text == 'Назад':
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            btnScheduleForTomorrow = types.KeyboardButton('Расписание на завтра')
            btnSheduleForTodayy = types.KeyboardButton('Расписание на сегодня')
            btnSheduleForDate = types.KeyboardButton('Расписание на дату')
            btnUpdateTable = types.KeyboardButton('Обновить таблицу')
            btnEditGroup = types.KeyboardButton('Изменить группу')
            btnError = types.KeyboardButton('Сообщить об ошибке')
            btnDonat = types.KeyboardButton('Пожертвование')
            markup.add(btnScheduleForTomorrow, btnSheduleForTodayy, btnSheduleForDate, btnUpdateTable, btnEditGroup, btnError, btnDonat)
            bot.send_message(message.chat.id,"Вы находитесь в меню", reply_markup=markup)
        if message.text == 'Меню':
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            btnScheduleForTomorrow = types.KeyboardButton('Расписание на завтра')
            btnSheduleForTodayy = types.KeyboardButton('Расписание на сегодня')
            btnSheduleForDate = types.KeyboardButton('Расписание на дату')
            btnUpdateTable = types.KeyboardButton('Обновить таблицу')
            btnEditGroup = types.KeyboardButton('Изменить группу')
            btnError = types.KeyboardButton('Сообщить об ошибке')
            btnDonat = types.KeyboardButton('Пожертвование')
            markup.add(btnScheduleForTomorrow, btnSheduleForTodayy, btnSheduleForDate, btnUpdateTable, btnEditGroup, btnError, btnDonat)
            bot.send_message(message.chat.id,"Вы находитесь в меню", reply_markup=markup)
        if message.text == 'Сообщить об ошибке':
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            btnScheduleForTomorrow = types.KeyboardButton('Расписание на завтра')
            btnSheduleForTodayy = types.KeyboardButton('Расписание на сегодня')
            btnSheduleForDate = types.KeyboardButton('Расписание на дату')
            btnUpdateTable = types.KeyboardButton('Обновить таблицу')
            btnEditGroup = types.KeyboardButton('Изменить группу')
            btnError = types.KeyboardButton('Сообщить об ошибке')
            btnDonat = types.KeyboardButton('Пожертвование')
            markup.add(btnScheduleForTomorrow, btnSheduleForTodayy, btnSheduleForDate, btnUpdateTable, btnEditGroup, btnError, btnDonat)
            bot.send_message(message.chat.id,"Напиши в бот:\nОшибка (ваш текст)\nПример:\nОшибка не выводит расписание на завтра", reply_markup=markup)
    if message.text == '04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).1':
        user_groupid = 1
    elif message.text == '04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).2':
        user_groupid = 2
    elif message.text == '04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).4':
        user_groupid = 3
    elif message.text == '04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).5':
        user_groupid = 4
    elif message.text == '04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).6':
        user_groupid = 5
    elif message.text == '04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).1':
        user_groupid = 6
    elif message.text == '04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).2':
        user_groupid = 7
    elif message.text == '04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).4':
        user_groupid = 8
    elif message.text == '04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).5':
        user_groupid = 9
    elif message.text == '04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).1':
        user_groupid = 10
    elif message.text == '04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).2':
        user_groupid = 11
    elif message.text == '04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).3':
        user_groupid = 12
    elif message.text == '04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).4':
        user_groupid = 13
    elif message.text == '04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).5':
        user_groupid = 14
    else:
        user_groupid = 0
    if user_groupid != 0:
        user_id = message.chat.id
        groupload['{0}'.format(user_id)]['groupid'] = user_groupid
        with open('groups.json', 'w') as f:
            json.dump(groupload, f, indent=4, ensure_ascii=False)
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        btnScheduleForTomorrow = types.KeyboardButton('Расписание на завтра')
        btnSheduleForTodayy = types.KeyboardButton('Расписание на сегодня')
        btnSheduleForDate = types.KeyboardButton('Расписание на дату')
        btnUpdateTable = types.KeyboardButton('Обновить таблицу')
        btnEditGroup = types.KeyboardButton('Изменить группу')
        btnError = types.KeyboardButton('Сообщить об ошибке')
        btnDonat = types.KeyboardButton('Пожертвование')
        markup.add(btnScheduleForTomorrow, btnSheduleForTodayy, btnSheduleForDate, btnUpdateTable, btnEditGroup, btnError, btnDonat)
        bot.send_message(message.chat.id,"Данные сохранены", reply_markup=markup)
    if message.text == 'Расписание на завтра':
        data = (datetime.now() + timedelta(1)).strftime('%Y-%m-%d')
        excel_data_df = pandas.read_excel(
        'Gotov.xlsx',
        sheet_name='Лист1', skiprows=[0])
        with open('groups.json', 'r') as f:
            groupload = json.load(f)
        chatid = groupload['{0}'.format(message.chat.id)]["groupid"]
        chatid = chatid+2
        if chatid == 3:
            chatBuckv = "04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).1"
        if chatid == 4:
            chatBuckv = "04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).2"
        if chatid == 5:
            chatBuckv = "04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).4"
        if chatid == 6:
            chatBuckv = "04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).5"
        if chatid == 7:
            chatBuckv = "04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).6"
        if chatid == 8:
            chatBuckv = "04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).1"
        if chatid == 9:
            chatBuckv = "04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).2"
        if chatid == 10:
            chatBuckv = "04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).4"
        if chatid == 11:
            chatBuckv = "04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).5"
        if chatid == 12:
            chatBuckv = "04.ЛОБ.21.ИЯИКСИЯ(АЯИИЯ).1"
        if chatid == 13:
            chatBuckv = "04.ЛОБ.21.ИЯИКСИЯ(АЯИИЯ).2"
        if chatid == 14:
            chatBuckv = "04.ЛОБ.21.ИЯИКСИЯ(АЯИИЯ).3"
        if chatid == 15:
            chatBuckv = "04.ЛОБ.21.ИЯИКСИЯ(АЯИИЯ).4"
        if chatid == 16:
            chatBuckv = "04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).5"
        excel_data_df['День'] = excel_data_df['День'].fillna(method='ffill')
        massiv_index = excel_data_df.index [excel_data_df['День']== data ]. tolist()
        if massiv_index == []:
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            btnScheduleForTomorrow = types.KeyboardButton('Расписание на завтра')
            btnSheduleForTodayy = types.KeyboardButton('Расписание на сегодня')
            btnSheduleForDate = types.KeyboardButton('Расписание на дату')
            btnUpdateTable = types.KeyboardButton('Обновить таблицу')
            btnEditGroup = types.KeyboardButton('Изменить группу')
            btnError = types.KeyboardButton('Сообщить об ошибке')
            btnDonat = types.KeyboardButton('Пожертвование')
            markup.add(btnScheduleForTomorrow, btnSheduleForTodayy, btnSheduleForDate, btnUpdateTable, btnEditGroup, btnError, btnDonat)
            bot.send_message(message.chat.id,"Завтра выходной 🥳", reply_markup=markup)
        else:
            d1,d2,d3,d4,d5,d6,d7,d8 = excel_data_df['Пара '].loc[excel_data_df.index[massiv_index]]
            i1,i2,i3,i4,i5,i6,i7,i8 = excel_data_df[chatBuckv].loc[excel_data_df.index[massiv_index]]
            if type(i1) not in [str]:
                i1n = math.isnan(i1)
            elif type(i1) == str:
                i1n = False
            if type(i2) not in [str]:
                i2n = math.isnan(i2)
            elif type(i2) == str:
                i2n = False
            if type(i3) not in [str]:
                i3n = math.isnan(i3)
            elif type(i3) == str:
                i3n = False
            if type(i4) not in [str]:
                i4n = math.isnan(i4)
            elif type(i4) == str:
                i4n = False
            if type(i5) not in [str]:
                i5n = math.isnan(i5)
            elif type(i5) == str:
                i5n = False
            if type(i6) not in [str]:
                i6n = math.isnan(i6)
            elif type(i6) == str:
                i6n = False
            if type(i7) not in [str]:
                i7n = math.isnan(i7)
            elif type(i7) == str:
                i7n = False
            if type(i8) not in [str]:
                i8n = math.isnan(i8)
            elif type(i8) == str:
                i8n = False
            if i1n == True & i2n == True & i3n == True & i4n == True & i5n == True & i6n == True & i7n == True & i8n == True:
                markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
                btnScheduleForTomorrow = types.KeyboardButton('Расписание на завтра')
                btnSheduleForTodayy = types.KeyboardButton('Расписание на сегодня')
                btnSheduleForDate = types.KeyboardButton('Расписание на дату')
                btnUpdateTable = types.KeyboardButton('Обновить таблицу')
                btnEditGroup = types.KeyboardButton('Изменить группу')
                btnError = types.KeyboardButton('Сообщить об ошибке')
                btnDonat = types.KeyboardButton('Пожертвование')
                markup.add(btnScheduleForTomorrow, btnSheduleForTodayy, btnSheduleForDate, btnUpdateTable, btnEditGroup, btnError, btnDonat)
                bot.send_message(message.chat.id,"Завтра выходной 🥳", reply_markup=markup)
            else:
                if type(i1) not in [str]:
                    i1 = 'Нет пары'
                if type(i2) not in [str]:
                    i2 = 'Нет пары'
                if type(i3) not in [str]:
                    i3 = 'Нет пары'
                if type(i4) not in [str]:
                    i4 = 'Нет пары'
                if type(i5) not in [str]:
                    i5 = 'Нет пары'
                if type(i6) not in [str]:
                    i6 = 'Нет пары'
                if type(i7) not in [str]:
                    i7 = 'Нет пары'
                if type(i8) not in [str]:
                    i8 = 'Нет пары'
                markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
                btnScheduleForTomorrow = types.KeyboardButton('Расписание на завтра')
                btnSheduleForTodayy = types.KeyboardButton('Расписание на сегодня')
                btnSheduleForDate = types.KeyboardButton('Расписание на дату')
                btnUpdateTable = types.KeyboardButton('Обновить таблицу')
                btnEditGroup = types.KeyboardButton('Изменить группу')
                btnError = types.KeyboardButton('Сообщить об ошибке')
                btnDonat = types.KeyboardButton('Пожертвование')
                markup.add(btnScheduleForTomorrow, btnSheduleForTodayy, btnSheduleForDate, btnUpdateTable, btnEditGroup, btnError, btnDonat)
                bot.send_message(message.chat.id,'{0}\n{1}\n\n{2}\n{3}\n\n{4}\n{5}\n\n{6}\n{7}\n\n{8}\n{9}\n\n{10}\n{11}\n\n{12}\n{13}\n\n{14}\n{15}'
                .format(d1,i1,d2,i2,d3,i3,d4,i4,d5,i5,d6,i6,d7,i7,d8,i8), reply_markup=markup)

    if message.text == 'Расписание на сегодня':
        data = (datetime.now()).strftime('%Y-%m-%d')
        excel_data_df = pandas.read_excel(
        'Gotov.xlsx',
        sheet_name='Лист1', skiprows=[0])
        with open('groups.json', 'r') as f:
            groupload = json.load(f)
        chatid = groupload['{0}'.format(message.chat.id)]["groupid"]
        chatid = chatid+2
        if chatid == 3:
            chatBuckv = "04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).1"
        if chatid == 4:
            chatBuckv = "04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).2"
        if chatid == 5:
            chatBuckv = "04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).4"
        if chatid == 6:
            chatBuckv = "04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).5"
        if chatid == 7:
            chatBuckv = "04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).6"
        if chatid == 8:
            chatBuckv = "04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).1"
        if chatid == 9:
            chatBuckv = "04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).2"
        if chatid == 10:
            chatBuckv = "04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).4"
        if chatid == 11:
            chatBuckv = "04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).5"
        if chatid == 12:
            chatBuckv = "04.ЛОБ.21.ИЯИКСИЯ(АЯИИЯ).1"
        if chatid == 13:
            chatBuckv = "04.ЛОБ.21.ИЯИКСИЯ(АЯИИЯ).2"
        if chatid == 14:
            chatBuckv = "04.ЛОБ.21.ИЯИКСИЯ(АЯИИЯ).3"
        if chatid == 15:
            chatBuckv = "04.ЛОБ.21.ИЯИКСИЯ(АЯИИЯ).4"
        if chatid == 16:
            chatBuckv = "04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).5"
        excel_data_df['День'] = excel_data_df['День'].fillna(method='ffill')
        massiv_index = excel_data_df.index [excel_data_df['День']== data ]. tolist()
        if massiv_index == []:
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            btnScheduleForTomorrow = types.KeyboardButton('Расписание на завтра')
            btnSheduleForTodayy = types.KeyboardButton('Расписание на сегодня')
            btnSheduleForDate = types.KeyboardButton('Расписание на дату')
            btnUpdateTable = types.KeyboardButton('Обновить таблицу')
            btnEditGroup = types.KeyboardButton('Изменить группу')
            btnError = types.KeyboardButton('Сообщить об ошибке')
            btnDonat = types.KeyboardButton('Пожертвование')
            markup.add(btnScheduleForTomorrow, btnSheduleForTodayy, btnSheduleForDate, btnUpdateTable, btnEditGroup, btnError, btnDonat)
            bot.send_message(message.chat.id,"Сегодня выходной 🥳", reply_markup=markup)
        else:
            d1,d2,d3,d4,d5,d6,d7,d8 = excel_data_df['Пара '].loc[excel_data_df.index[massiv_index]]
            i1,i2,i3,i4,i5,i6,i7,i8 = excel_data_df[chatBuckv].loc[excel_data_df.index[massiv_index]]
            if type(i1) not in [str]:
                i1n = math.isnan(i1)
            elif type(i1) == str:
                i1n = False
            if type(i2) not in [str]:
                i2n = math.isnan(i2)
            elif type(i2) == str:
                i2n = False
            if type(i3) not in [str]:
                i3n = math.isnan(i3)
            elif type(i3) == str:
                i3n = False
            if type(i4) not in [str]:
                i4n = math.isnan(i4)
            elif type(i4) == str:
                i4n = False
            if type(i5) not in [str]:
                i5n = math.isnan(i5)
            elif type(i5) == str:
                i5n = False
            if type(i6) not in [str]:
                i6n = math.isnan(i6)
            elif type(i6) == str:
                i6n = False
            if type(i7) not in [str]:
                i7n = math.isnan(i7)
            elif type(i7) == str:
                i7n = False
            if type(i8) not in [str]:
                i8n = math.isnan(i8)
            elif type(i8) == str:
                i8n = False
            if i1n == True & i2n == True & i3n == True & i4n == True & i5n == True & i6n == True & i7n == True & i8n == True:
                markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
                btnScheduleForTomorrow = types.KeyboardButton('Расписание на завтра')
                btnSheduleForTodayy = types.KeyboardButton('Расписание на сегодня')
                btnSheduleForDate = types.KeyboardButton('Расписание на дату')
                btnUpdateTable = types.KeyboardButton('Обновить таблицу')
                btnEditGroup = types.KeyboardButton('Изменить группу')
                btnError = types.KeyboardButton('Сообщить об ошибке')
                btnDonat = types.KeyboardButton('Пожертвование')
                markup.add(btnScheduleForTomorrow, btnSheduleForTodayy, btnSheduleForDate, btnUpdateTable, btnEditGroup, btnError, btnDonat)
                bot.send_message(message.chat.id,"Сегодня выходной 🥳", reply_markup=markup)
            else:
                if type(i1) not in [str]:
                    i1 = 'Нет пары'
                if type(i2) not in [str]:
                    i2 = 'Нет пары'
                if type(i3) not in [str]:
                    i3 = 'Нет пары'
                if type(i4) not in [str]:
                    i4 = 'Нет пары'
                if type(i5) not in [str]:
                    i5 = 'Нет пары'
                if type(i6) not in [str]:
                    i6 = 'Нет пары'
                if type(i7) not in [str]:
                    i7 = 'Нет пары'
                if type(i8) not in [str]:
                    i8 = 'Нет пары'
                markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
                btnScheduleForTomorrow = types.KeyboardButton('Расписание на завтра')
                btnSheduleForTodayy = types.KeyboardButton('Расписание на сегодня')
                btnSheduleForDate = types.KeyboardButton('Расписание на дату')
                btnUpdateTable = types.KeyboardButton('Обновить таблицу')
                btnEditGroup = types.KeyboardButton('Изменить группу')
                btnError = types.KeyboardButton('Сообщить об ошибке')
                btnDonat = types.KeyboardButton('Пожертвование')
                markup.add(btnScheduleForTomorrow, btnSheduleForTodayy, btnSheduleForDate, btnUpdateTable, btnEditGroup, btnError, btnDonat)
                bot.send_message(message.chat.id,'{0}\n{1}\n\n{2}\n{3}\n\n{4}\n{5}\n\n{6}\n{7}\n\n{8}\n{9}\n\n{10}\n{11}\n\n{12}\n{13}\n\n{14}\n{15}'
                .format(d1,i1,d2,i2,d3,i3,d4,i4,d5,i5,d6,i6,d7,i7,d8,i8), reply_markup=markup)
    if message.text == f'админ':
        if str(message.chat.id) == '615009766':
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            btnMenu = types.KeyboardButton('Меню')
            markup.add(btnMenu)
            bot.send_message(message.chat.id,'Готово', reply_markup=markup)

    obnov = [f'обновление']
    for obnovs in obnov:
        if obnovs in message.text.lower():
            if str(message.chat.id) == '615009766':
                with open('groups.json', 'r') as f:
                    groupload = json.load(f)    
                orgs = [x for x in groupload]
                for i in orgs:
                    bot.send_message(i,'{0}'.format(message.text))
        
bot.infinity_polling()