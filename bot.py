import pandas
import json
import math
import config
import time 
import asyncio
import aioschedule
import requests
import os
import datetime as dt
from datetime import datetime 
from aiogram import Bot, Dispatcher, executor, types
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove, InlineKeyboardButton, InlineKeyboardMarkup
from aiogram.dispatcher.filters import Text
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.dispatcher.filters.state import StatesGroup, State
from aiogram.dispatcher import FSMContext
from datetime import timedelta
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

bot = Bot(config.token)
dp = Dispatcher(bot, storage=MemoryStorage())

storage = MemoryStorage()

os.chdir(r'C:\Users\strai\source\repos\VSK\MGOPA')

class ProfileStatesGroup(StatesGroup):
    date = State()
    error = State()
    yesnou = State()
    file = State()
    usernamepol = State()
    message = State()
    yesnoumessage = State()
    messageall = State()
    yesnoumessageall = State()

ikb = InlineKeyboardMarkup(resize_keyboard=True, row_width=1)
ibtn1 = InlineKeyboardButton('04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).1',callback_data='04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).1')
ibtn2 = InlineKeyboardButton('04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).2',callback_data='04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).2')
ibtn3 = InlineKeyboardButton('04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).4',callback_data='04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).4')
ibtn4 = InlineKeyboardButton('04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).5',callback_data='04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).5')
ibtn5 = InlineKeyboardButton('04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).6',callback_data='04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).6')
ibtn6 = InlineKeyboardButton('04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).1',callback_data='04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).1')
ibtn7 = InlineKeyboardButton('04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).2',callback_data='04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).2')
ibtn8 = InlineKeyboardButton('04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).4',callback_data='04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).4')
ibtn9 = InlineKeyboardButton('04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).5',callback_data='04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).5')
ibtn10 = InlineKeyboardButton('04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).1',callback_data='04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).1')
ibtn11 = InlineKeyboardButton('04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).2',callback_data='04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).2')
ibtn12 = InlineKeyboardButton('04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).3',callback_data='04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).3')
ibtn13 = InlineKeyboardButton('04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).4',callback_data='04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).4')
ibtn14 = InlineKeyboardButton('04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).5',callback_data='04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).5')
InlineGroups = ikb.add(ibtn1,ibtn2,ibtn3,ibtn4,ibtn5,ibtn6,ibtn7,ibtn8,ibtn9,ibtn10,ibtn11,ibtn12,ibtn13,ibtn14)

ikb1 = InlineKeyboardMarkup(resize_keyboard=True, row_width=1)
ibtnmenu = InlineKeyboardButton('Меню',callback_data='Меню')
InlineGroupsAndMenu = ikb1.add(ibtn1,ibtn2,ibtn3,ibtn4,ibtn5,ibtn6,ibtn7,ibtn8,ibtn9,ibtn10,ibtn11,ibtn12,ibtn13,ibtn14,ibtnmenu)

kbBack = ReplyKeyboardMarkup(resize_keyboard=True)
btnBack = KeyboardButton('Назад')
kbBack = kbBack.add(btnBack)
                     
kb = ReplyKeyboardMarkup(resize_keyboard=True)
btnScheduleForTomorrow = KeyboardButton('Расписание на завтра')
btnSheduleForTodayy = KeyboardButton('Расписание на сегодня')
btnSheduleForDate = KeyboardButton('Расписание на дату')
btnUpdateTable = KeyboardButton('Обновить таблицу')
btnEditGroup = KeyboardButton('Изменить группу')
btnError = KeyboardButton('Сообщить об ошибке')
btnDonat = KeyboardButton('Пожертвование')
MainButtons = kb.add(btnScheduleForTomorrow,
                     btnSheduleForTodayy).insert(btnSheduleForDate).add(btnUpdateTable,
                                                                        btnEditGroup).insert(btnError).add(btnDonat)

kbAdmin = ReplyKeyboardMarkup(resize_keyboard=True)
btnScheduleAdmin = KeyboardButton('Загрузить расписание')
btnListUserAdmin = KeyboardButton('Список пользователей')
btnListAdmin = KeyboardButton('Список админов')
btnMenuAdmin = KeyboardButton('Меню')
btnMessageAdmin = KeyboardButton('Написать')
btnMessageAllAdmin = KeyboardButton('Написать всем')
AdminMenu = kbAdmin.add(btnListUserAdmin,btnListAdmin, btnMessageAdmin, btnMessageAllAdmin, btnMenuAdmin, btnScheduleAdmin)

kbpusto = ReplyKeyboardMarkup(resize_keyboard=True)
kbpusto = kbpusto.add()

kbyesnou = ReplyKeyboardMarkup(resize_keyboard=True)
btnYes = KeyboardButton('Да')
btnNot = KeyboardButton('Нет')
kbyesnou = kbyesnou.add(btnYes,btnNot)

async def on_startup(_):
    print('Бот запушен')
    asyncio.create_task(scheduler())

async def scheduler():
    aioschedule.every().day.at("00:00").do(upgrade)
    while True:
        await aioschedule.run_pending()
        await asyncio.sleep(1)        

async def save_user(user_groupid,user_id,user_username):
    with open('groups.json', 'r') as f:
        groupload = json.load(f)
    if str(user_id) not in groupload:
        if user_groupid != 0:
            with open('groups.json', 'w') as f:
                groupload[user_id] = {'groupid': user_groupid,'username':user_username}
                json.dump(groupload, f, indent=4, ensure_ascii=False)
                await bot.send_message(user_id, text='Данные сохранены', reply_markup=MainButtons)
    else:
        if user_groupid != 0:
            groupload[f'{user_id}']['groupid'] = user_groupid
            with open('groups.json', 'w') as f:
                json.dump(groupload, f, indent=4, ensure_ascii=False)
            groupload[f'{user_id}']['username'] = user_username
            with open('groups.json', 'w') as f:
                json.dump(groupload, f, indent=4, ensure_ascii=False)
            await bot.send_message(user_id, text='Данные сохранены', reply_markup=MainButtons)
        elif user_groupid == 0:
            groupload[f'{user_id}']['username'] = user_username
            with open('groups.json', 'w') as f:
                json.dump(groupload, f, indent=4, ensure_ascii=False)
            await bot.send_message(user_id, text='Меню', reply_markup=MainButtons)

async def upgrade():
    with open('url.json', 'r') as f:
        urlload = json.load(f)
    url = urlload["url"]["url"]
    resp = requests.get(url)
    output = open('45_03_02_Иностранные_языки_и_культуры_стран_изучаемых_языков_английский.xlsx', 'wb')
    output.write(resp.content)
    output.close()
    wb = load_workbook('45_03_02_Иностранные_языки_и_культуры_стран_изучаемых_языков_английский.xlsx')
    sheet_ranges = wb['Лист1']
    bokv = sheet_ranges.merged_cells.ranges
    for st_name in wb.sheetnames:
        st = wb[st_name]
        mcr_coord_list = [mcr.coord for mcr in st.merged_cells.ranges]
        
        for mcr in mcr_coord_list:
            min_col, min_row, max_col, max_row = range_boundaries(mcr)
            top_left_cell_value = st.cell(row=min_row, column=min_col).value
            st.unmerge_cells(mcr)
            for row in st.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
                for cell in row:
                    cell.value = top_left_cell_value
    wb.save('Gotov.xlsx')


@dp.message_handler(commands=['start'])
async def start_command(message: types.Message):
    with open('groups.json', 'r') as f:
        groupload = json.load(f)
    
    if str(message.chat.id) not in groupload:
        await message.answer(text=f'<em>Привет {message.chat.full_name}, я МГОПУ бот🤖\nВыбери свою группу</em>', parse_mode='HTML', reply_markup=InlineGroups)
        await message.delete()
    else:
        await message.answer(text=f'<em>Привет {message.chat.full_name}, я МГОПУ бот🤖\nВыбери свою группу</em>', parse_mode='HTML', reply_markup=InlineGroupsAndMenu)
        await message.delete()

@dp.message_handler(commands=['update'])
async def upgrade_command(message: types.Message):
    await message.answer(text='Обновление..')
    await upgrade()
    await message.answer(text='Таблица обновлена')

@dp.message_handler(commands=['message'])
async def error_command(message: types.Message):
    await message.answer(text='Напишите что у вас случилось', reply_markup=kbBack)
    await ProfileStatesGroup.error.set()
    

@dp.message_handler(commands=['admin'])
async def admin_command(message: types.Message):
    await message.answer(text='Админ панель', reply_markup=AdminMenu)
    await message.delete()

@dp.callback_query_handler()
async def groups_callback(callback: types.CallbackQuery):
    user_id = callback.from_user.id
    user_username = callback.from_user.username
    if callback.data == '04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).1':
        user_groupid = 1
        await save_user(user_groupid,user_id,user_username)
        await callback.message.delete()
    elif callback.data == '04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).2':
        user_groupid = 2
        await save_user(user_groupid,user_id,user_username)
        await callback.message.delete()
    elif callback.data == '04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).4':
        user_groupid = 3
        await save_user(user_groupid,user_id,user_username)
        await callback.message.delete()
    elif callback.data == '04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).5':
        user_groupid = 4
        await save_user(user_groupid,user_id,user_username)
        await callback.message.delete()
    elif callback.data == '04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).6':
        user_groupid = 5
        await save_user(user_groupid,user_id,user_username)
        await callback.message.delete()
    elif callback.data == '04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).1':
        user_groupid = 6
        await save_user(user_groupid,user_id,user_username)
        await callback.message.delete()
    elif callback.data == '04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).2':
        user_groupid = 7
        await save_user(user_groupid,user_id,user_username)
        await callback.message.delete()
    elif callback.data == '04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).4':
        user_groupid = 8
        await save_user(user_groupid,user_id,user_username)
        await callback.message.delete()
    elif callback.data == '04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).5':
        user_groupid = 9
        await save_user(user_groupid,user_id,user_username)
        await callback.message.delete()
    elif callback.data == '04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).1':
        user_groupid = 10
        await save_user(user_groupid,user_id,user_username)
        await callback.message.delete()
    elif callback.data == '04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).2':
        user_groupid = 11
        await save_user(user_groupid,user_id,user_username)
        await callback.message.delete()
    elif callback.data == '04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).3':
        user_groupid = 12
        await save_user(user_groupid,user_id,user_username)
        await callback.message.delete()
    elif callback.data == '04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).4':
        user_groupid = 13
        await save_user(user_groupid,user_id,user_username)
        await callback.message.delete()
    elif callback.data == '04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).5':
        user_groupid = 14
        await save_user(user_groupid,user_id,user_username)
        await callback.message.delete()
    elif callback.data == 'Меню':
        user_groupid = 0
        await save_user(user_groupid,user_id,user_username)
        await callback.message.delete()

@dp.message_handler(Text)
async def echo(message: types.Message):
    user_id = message.from_user.id
    with open('groups.json', 'r') as f:
        groupload = json.load(f)
    if str(user_id) not in groupload:
        await message.answer(text=f'<em>Привет {message.chat.full_name}, я МГОПУ бот🤖\nВыбери свою группу</em>',
                              parse_mode='HTML', reply_markup=InlineGroups)
        await message.delete()
    else:
        if message.text == 'Обновить таблицу':
            await message.answer(text=f'Обновлять таблицу можно только в том случае, если у вас не показывается <b>долго</b> расписание\n'+
                                 'Тогда напишите или нажмите на эту команду -> /update',parse_mode='HTML', reply_markup=MainButtons)
        elif message.text == 'Пожертвование':
            with open('card.json', 'r') as af:
                card = json.load(af)
            await message.answer(text=f'Сбер: {card["Sber"]}\nТинькофф: {card["Tinkoff"]}', reply_markup=MainButtons)
        elif message.text == 'Изменить группу':
            msg = await message.answer(text='Выберите группу', reply_markup=ReplyKeyboardRemove())
            next_id = msg.message_id
            await bot.delete_message(message.from_user.id, next_id)
            await message.answer(text='Выберите группу', reply_markup=InlineGroupsAndMenu)
            await message.delete()
        elif message.text == 'Расписание на дату':
            await message.answer(text='Напиши дату на которое нужно узнать расписание', reply_markup=kbBack)
            await ProfileStatesGroup.date.set()
        elif message.text == 'Сообщить об ошибке':
            await message.answer(text='Напишите или нажмите на команду -> /message\nИ напишите, в чем проблема', reply_markup=MainButtons)
        elif message.text == 'Расписание на завтра':
            offset = dt.timezone(dt.timedelta(hours=3))
            data = (dt.datetime.now(offset) + timedelta(1)).strftime('%Y-%m-%d')
            print(data)
            # data = (datetime.now() + timedelta(1)).strftime('%Y-%m-%d')
            excel_data_df = pandas.read_excel(
            'Gotov.xlsx',
            sheet_name='Лист1', skiprows=[0])
            with open('groups.json', 'r') as f:
                groupload = json.load(f)
            chatid = groupload['{0}'.format(message.chat.id)]["groupid"]
            chatid = chatid+2
            if chatid == 3:
                chatBuckv = "04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).1"
            if chatid == 4:
                chatBuckv = "04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).2"
            if chatid == 5:
                chatBuckv = "04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).4"
            if chatid == 6:
                chatBuckv = "04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).5"
            if chatid == 7:
                chatBuckv = "04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).6"
            if chatid == 8:
                chatBuckv = "04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).1"
            if chatid == 9:
                chatBuckv = "04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).2"
            if chatid == 10:
                chatBuckv = "04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).4"
            if chatid == 11:
                chatBuckv = "04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).5"
            if chatid == 12:
                chatBuckv = "04.ЛОБ.21.ИЯИКСИЯ(АЯИИЯ).1"
            if chatid == 13:
                chatBuckv = "04.ЛОБ.21.ИЯИКСИЯ(АЯИИЯ).2"
            if chatid == 14:
                chatBuckv = "04.ЛОБ.21.ИЯИКСИЯ(АЯИИЯ).3"
            if chatid == 15:
                chatBuckv = "04.ЛОБ.21.ИЯИКСИЯ(АЯИИЯ).4"
            if chatid == 16:
                chatBuckv = "04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).5"
            excel_data_df['День'] = excel_data_df['День'].fillna(method='ffill')
            massiv_index = excel_data_df.index [excel_data_df['День']== data ]. tolist()
            if massiv_index == []:
                await message.answer(text='Завтра выходной 🥳', reply_markup=MainButtons)
            else:
                d1,d2,d3,d4,d5,d6,d7,d8 = excel_data_df['Пара '].loc[excel_data_df.index[massiv_index]]
                i1,i2,i3,i4,i5,i6,i7,i8 = excel_data_df[chatBuckv].loc[excel_data_df.index[massiv_index]]
                if type(i1) not in [str]:
                    i1n = math.isnan(i1)
                elif type(i1) == str:
                    i1n = False
                if type(i2) not in [str]:
                    i2n = math.isnan(i2)
                elif type(i2) == str:
                    i2n = False
                if type(i3) not in [str]:
                    i3n = math.isnan(i3)
                elif type(i3) == str:
                    i3n = False
                if type(i4) not in [str]:
                    i4n = math.isnan(i4)
                elif type(i4) == str:
                    i4n = False
                if type(i5) not in [str]:
                    i5n = math.isnan(i5)
                elif type(i5) == str:
                    i5n = False
                if type(i6) not in [str]:
                    i6n = math.isnan(i6)
                elif type(i6) == str:
                    i6n = False
                if type(i7) not in [str]:
                    i7n = math.isnan(i7)
                elif type(i7) == str:
                    i7n = False
                if type(i8) not in [str]:
                    i8n = math.isnan(i8)
                elif type(i8) == str:
                    i8n = False
                if i1n == True & i2n == True & i3n == True & i4n == True & i5n == True & i6n == True & i7n == True & i8n == True:
                    await message.answer(text='Завтра выходной 🥳', reply_markup=MainButtons)
                else:
                    if type(i1) not in [str]:
                        i1 = 'Нет пары'
                    if type(i2) not in [str]:
                        i2 = 'Нет пары'
                    if type(i3) not in [str]:
                        i3 = 'Нет пары'
                    if type(i4) not in [str]:
                        i4 = 'Нет пары'
                    if type(i5) not in [str]:
                        i5 = 'Нет пары'
                    if type(i6) not in [str]:
                        i6 = 'Нет пары'
                    if type(i7) not in [str]:
                        i7 = 'Нет пары'
                    if type(i8) not in [str]:
                        i8 = 'Нет пары'
                    await message.answer(text='{0}\n{1}\n\n{2}\n{3}\n\n{4}\n{5}\n\n{6}\n{7}\n\n{8}\n{9}\n\n{10}\n{11}\n\n{12}\n{13}\n\n{14}\n{15}'
                    .format(d1,i1,d2,i2,d3,i3,d4,i4,d5,i5,d6,i6,d7,i7,d8,i8), reply_markup=MainButtons)
        elif message.text == 'Расписание на сегодня':
            offset = dt.timezone(dt.timedelta(hours=3))
            data = dt.datetime.now(offset).strftime('%Y-%m-%d')
            print(data)
            # data = (datetime.now()).strftime('%Y-%m-%d')
            excel_data_df = pandas.read_excel(
            'Gotov.xlsx',
            sheet_name='Лист1', skiprows=[0])
            with open('groups.json', 'r') as f:
                groupload = json.load(f)
            chatid = groupload['{0}'.format(message.chat.id)]["groupid"]
            chatid = chatid+2
            if chatid == 3:
                chatBuckv = "04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).1"
            if chatid == 4:
                chatBuckv = "04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).2"
            if chatid == 5:
                chatBuckv = "04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).4"
            if chatid == 6:
                chatBuckv = "04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).5"
            if chatid == 7:
                chatBuckv = "04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).6"
            if chatid == 8:
                chatBuckv = "04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).1"
            if chatid == 9:
                chatBuckv = "04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).2"
            if chatid == 10:
                chatBuckv = "04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).4"
            if chatid == 11:
                chatBuckv = "04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).5"
            if chatid == 12:
                chatBuckv = "04.ЛОБ.21.ИЯИКСИЯ(АЯИИЯ).1"
            if chatid == 13:
                chatBuckv = "04.ЛОБ.21.ИЯИКСИЯ(АЯИИЯ).2"
            if chatid == 14:
                chatBuckv = "04.ЛОБ.21.ИЯИКСИЯ(АЯИИЯ).3"
            if chatid == 15:
                chatBuckv = "04.ЛОБ.21.ИЯИКСИЯ(АЯИИЯ).4"
            if chatid == 16:
                chatBuckv = "04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).5"
            excel_data_df['День'] = excel_data_df['День'].fillna(method='ffill')
            massiv_index = excel_data_df.index [excel_data_df['День']== data ]. tolist()
            if massiv_index == []:
                await message.answer(text='Сегодня выходной 🥳', reply_markup=MainButtons)
            else:
                d1,d2,d3,d4,d5,d6,d7,d8 = excel_data_df['Пара '].loc[excel_data_df.index[massiv_index]]
                i1,i2,i3,i4,i5,i6,i7,i8 = excel_data_df[chatBuckv].loc[excel_data_df.index[massiv_index]]
                if type(i1) not in [str]:
                    i1n = math.isnan(i1)
                elif type(i1) == str:
                    i1n = False
                if type(i2) not in [str]:
                    i2n = math.isnan(i2)
                elif type(i2) == str:
                    i2n = False
                if type(i3) not in [str]:
                    i3n = math.isnan(i3)
                elif type(i3) == str:
                    i3n = False
                if type(i4) not in [str]:
                    i4n = math.isnan(i4)
                elif type(i4) == str:
                    i4n = False
                if type(i5) not in [str]:
                    i5n = math.isnan(i5)
                elif type(i5) == str:
                    i5n = False
                if type(i6) not in [str]:
                    i6n = math.isnan(i6)
                elif type(i6) == str:
                    i6n = False
                if type(i7) not in [str]:
                    i7n = math.isnan(i7)
                elif type(i7) == str:
                    i7n = False
                if type(i8) not in [str]:
                    i8n = math.isnan(i8)
                elif type(i8) == str:
                    i8n = False
                if i1n == True & i2n == True & i3n == True & i4n == True & i5n == True & i6n == True & i7n == True & i8n == True:
                    await message.answer(text='Сегодня выходной 🥳', reply_markup=MainButtons)
                else:
                    if type(i1) not in [str]:
                        i1 = 'Нет пары'
                    if type(i2) not in [str]:
                        i2 = 'Нет пары'
                    if type(i3) not in [str]:
                        i3 = 'Нет пары'
                    if type(i4) not in [str]:
                        i4 = 'Нет пары'
                    if type(i5) not in [str]:
                        i5 = 'Нет пары'
                    if type(i6) not in [str]:
                        i6 = 'Нет пары'
                    if type(i7) not in [str]:
                        i7 = 'Нет пары'
                    if type(i8) not in [str]:
                        i8 = 'Нет пары'
                    await message.answer(text='{0}\n{1}\n\n{2}\n{3}\n\n{4}\n{5}\n\n{6}\n{7}\n\n{8}\n{9}\n\n{10}\n{11}\n\n{12}\n{13}\n\n{14}\n{15}'
                    .format(d1,i1,d2,i2,d3,i3,d4,i4,d5,i5,d6,i6,d7,i7,d8,i8), reply_markup=MainButtons)
        elif message.text == 'Меню':
            await message.answer(text='Меню', reply_markup=MainButtons)
            await message.delete()
        elif message.text == 'Загрузить расписание':
            with open('admin.json', 'r') as ad:
                admin = json.load(ad)
                if str(message.from_user.id) in admin:
                    await message.answer(text='Пришлите ссылку на рассписания', reply_markup=kbBack)
                    await message.delete()
                    await ProfileStatesGroup.file.set()
        elif message.text == 'Список пользователей':
            with open('admin.json', 'r') as ad:
                admin = json.load(ad)
                if str(message.from_user.id) in admin:
                    with open('groups.json', 'r') as f:
                        groupload = json.load(f)
                        for id in groupload.keys():
                            await bot.send_message(message.from_user.id, text=f'Id: {id}\nUsername: {groupload[id]["username"]}\nGroupid: {groupload[id]["groupid"]}')
                    await message.delete()
        elif message.text == 'Список админов':
            with open('admin.json', 'r') as ad:
                admin = json.load(ad)
                if str(message.from_user.id) in admin:
                    with open('admin.json', 'r') as f:
                        admines = json.load(f)
                        for id in admines.keys():
                            await bot.send_message(message.from_user.id, text=f'Id: {id}\nUsername: {admines[id]["username"]}')
                    await message.delete()
        elif message.text == 'Написать':
            with open('admin.json', 'r') as ad:
                admin = json.load(ad)
                if str(message.from_user.id) in admin:
                    await message.answer(text='Напишите username кому хотите написать', reply_markup=kbBack)
                    await ProfileStatesGroup.usernamepol.set()
                    await message.delete()
        elif message.text == 'Написать всем':
            with open('admin.json', 'r') as ad:
                admin = json.load(ad)
                if str(message.from_user.id) in admin:
                    await message.answer(text='Напишите сообщение', reply_markup=kbBack)
                    await ProfileStatesGroup.messageall.set()
                    await message.delete()

@dp.message_handler(state=ProfileStatesGroup.error)
async def load_error(message: types.Message, state: FSMContext) -> None:
    if message.text == 'Назад':
        await message.answer(text='Меню', reply_markup=MainButtons)
        await state.finish()
    else:
        async with state.proxy() as errors:
            errors['errors'] = message.text
        await message.answer(text='Вы уверены', reply_markup=kbyesnou)
        await ProfileStatesGroup.yesnou.set()

@dp.message_handler(state=ProfileStatesGroup.usernamepol)
async def load_error(message: types.Message, state: FSMContext) -> None:
    with open('groups.json', 'r') as f:
        groupload = json.load(f)
        username = message.text
        nulin = 0
        for id in groupload.keys():
            if groupload[id]["username"] == username:
                nulin = 1
                async with state.proxy() as datemessage:
                    datemessage['id'] = id
                await message.answer(text='Напишите сообщения для этого пользователя')
                await ProfileStatesGroup.message.set()
        if nulin == 0:
            await message.answer(text='Пользователь не найден')
            await ProfileStatesGroup.usernamepol.set()
        if message.text == 'Назад':
            await message.answer(text='Админ меню', reply_markup=AdminMenu)
            await state.finish()

@dp.message_handler(state=ProfileStatesGroup.messageall)
async def load_error(message: types.Message, state: FSMContext) -> None:
    if message.text == 'Назад':
        await message.answer(text='Админ меню', reply_markup=AdminMenu)
        await state.finish()
    else:
        async with state.proxy() as datemessage:
            datemessage['message'] = message.text
        await message.answer(text='Вы уверены?', reply_markup=kbyesnou)
        await ProfileStatesGroup.yesnoumessageall.set()

@dp.message_handler(state=ProfileStatesGroup.message)
async def load_error(message: types.Message, state: FSMContext) -> None:
    if message.text == 'Назад':
        await message.answer(text='Админ меню', reply_markup=AdminMenu)
        await state.finish()
    else:
        async with state.proxy() as datemessage:
            datemessage['message'] = message.text
        await message.answer(text='Вы уверены?', reply_markup=kbyesnou)
        await ProfileStatesGroup.yesnoumessage.set()

@dp.message_handler(state=ProfileStatesGroup.yesnoumessageall)
async def load_error(message: types.Message, state: FSMContext) -> None:
    if message.text == 'Да':
        async with state.proxy() as datemessage:
            datemessage['yes'] = message.text
        with open('groups.json', 'r') as f:
            groupload = json.load(f)    
        orgs = [x for x in groupload]
        for i in orgs:
            await bot.send_message(i, text=f'Сообщение от администратора: {datemessage["message"]}')
        await state.finish()
    else:
        await message.answer(text='Админ меню', reply_markup=AdminMenu)
        await state.finish()

@dp.message_handler(state=ProfileStatesGroup.yesnoumessage)
async def load_error(message: types.Message, state: FSMContext) -> None:
    if message.text == 'Да':
        async with state.proxy() as datemessage:
            datemessage['yesnoumessage'] = message.text
        await bot.send_message(datemessage['id'], text=f'Сообщение от администратора: {datemessage["message"]}')
        await message.answer(text='Сообщение отправлено', reply_markup=AdminMenu)
        await state.finish()
    else:
        await message.answer(text='Админ меню', reply_markup=AdminMenu)
        await state.finish()

@dp.message_handler(state=ProfileStatesGroup.yesnou)
async def load_error(message: types.Message, state: FSMContext) -> None:
    if message.text == 'Да':
        async with state.proxy() as errors:
            errors['yesnou'] = message.text
        await bot.send_message('615009766', text=f'id: {message.from_user.id}\nusername: {message.from_user.username}\nText: {errors["errors"]}')
        await message.answer(text='Сообщение отправлено', reply_markup=MainButtons)
        await state.finish()
    else:
        await message.answer(text='Напишите что у вас случилось', reply_markup=kbBack)
        await ProfileStatesGroup.error.set()

@dp.message_handler(state=ProfileStatesGroup.file)
async def load_url(message: types.Message, state: FSMContext) -> None:
    try:
        urlAddress = message.text
        efx = 'export?format=xlsx'

        while urlAddress[-1] != '/':
            urlAddress = urlAddress[:-1]
        urlAddress = urlAddress + efx

        with open('url.json', 'r') as f:
                urlload = json.load(f)

        with open('url.json', 'w') as f:
                urlload['url'] = {'url': urlAddress}
                json.dump(urlload, f, indent=4, ensure_ascii=False)
        await message.answer(text='Успешно измененно', reply_markup=AdminMenu)
        await state.finish()
    except ValueError:
        await message.answer(text='Ошибка', reply_markup=AdminMenu)
        await state.finish()

@dp.message_handler(state=ProfileStatesGroup.date)
async def load_date(message: types.Message, state: FSMContext) -> None:
    try:
        datetime.strptime(message.text, '%d.%m.%Y')
        dateT = True
    except ValueError:
        dateT = False
    if dateT == True:
        async with state.proxy() as data:
            data["date"] = message.text
        date = message.text
        data = datetime.strptime(date, '%d.%m.%Y')
        data = data.strftime('%Y-%m-%d')
        excel_data_df = pandas.read_excel(
        'Gotov.xlsx',
        sheet_name='Лист1', skiprows=[0])
        with open('groups.json', 'r') as f:
            groupload = json.load(f)
        chatid = groupload['{0}'.format(message.chat.id)]["groupid"]
        chatid = chatid+2
        if chatid == 3:
            chatBuckv = "04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).1"
        if chatid == 4:
            chatBuckv = "04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).2"
        if chatid == 5:
            chatBuckv = "04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).4"
        if chatid == 6:
            chatBuckv = "04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).5"
        if chatid == 7:
            chatBuckv = "04.ЛОБ.19.ИЯИКСИЯ(АЯИИЯ).6"
        if chatid == 8:
            chatBuckv = "04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).1"
        if chatid == 9:
            chatBuckv = "04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).2"
        if chatid == 10:
            chatBuckv = "04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).4"
        if chatid == 11:
            chatBuckv = "04.ЛОБ.20.ИЯИКСИЯ(АЯИИЯ).5"
        if chatid == 12:
            chatBuckv = "04.ЛОБ.21.ИЯИКСИЯ(АЯИИЯ).1"
        if chatid == 13:
            chatBuckv = "04.ЛОБ.21.ИЯИКСИЯ(АЯИИЯ).2"
        if chatid == 14:
            chatBuckv = "04.ЛОБ.21.ИЯИКСИЯ(АЯИИЯ).3"
        if chatid == 15:
            chatBuckv = "04.ЛОБ.21.ИЯИКСИЯ(АЯИИЯ).4"
        if chatid == 16:
            chatBuckv = "04.ЛОБ.21.ИЯИКСИЯ(АЯиИЯ).5"
        excel_data_df['День'] = excel_data_df['День'].fillna(method='ffill')
        massiv_index = excel_data_df.index [excel_data_df['День']== data ]. tolist()
        if massiv_index == []:
            await message.answer(text='Не могу найти эту дату', reply_markup=MainButtons)
            await state.finish()
        else:
            d1,d2,d3,d4,d5,d6,d7,d8 = excel_data_df['Пара '].loc[excel_data_df.index[massiv_index]]
            i1,i2,i3,i4,i5,i6,i7,i8 = excel_data_df[chatBuckv].loc[excel_data_df.index[massiv_index]]
            raz = excel_data_df[chatBuckv].loc[excel_data_df.index[massiv_index]]
            if type(i1) not in [str]:
                i1n = math.isnan(i1)
            elif type(i1) == str:
                i1n = False
            if type(i2) not in [str]:
                i2n = math.isnan(i2)
            elif type(i2) == str:
                i2n = False
            if type(i3) not in [str]:
                i3n = math.isnan(i3)
            elif type(i3) == str:
                i3n = False
            if type(i4) not in [str]:
                i4n = math.isnan(i4)
            elif type(i4) == str:
                i4n = False
            if type(i5) not in [str]:
                i5n = math.isnan(i5)
            elif type(i5) == str:
                i5n = False
            if type(i6) not in [str]:
                i6n = math.isnan(i6)
            elif type(i6) == str:
                i6n = False
            if type(i7) not in [str]:
                i7n = math.isnan(i7)
            elif type(i7) == str:
                i7n = False
            if type(i8) not in [str]:
                i8n = math.isnan(i8)
            elif type(i8) == str:
                i8n = False
            if i1n == True & i2n == True & i3n == True & i4n == True & i5n == True & i6n == True & i7n == True & i8n == True:
                await message.answer(text='Выходной 🥳', reply_markup=MainButtons)
                await state.finish()
            else:
                if type(i1) not in [str]:
                    i1 = 'Нет пары'
                if type(i2) not in [str]:
                    i2 = 'Нет пары'
                if type(i3) not in [str]:
                    i3 = 'Нет пары'
                if type(i4) not in [str]:
                    i4 = 'Нет пары'
                if type(i5) not in [str]:
                    i5 = 'Нет пары'
                if type(i6) not in [str]:
                    i6 = 'Нет пары'
                if type(i7) not in [str]:
                    i7 = 'Нет пары'
                if type(i8) not in [str]:
                    i8 = 'Нет пары'
                await message.answer(text='{0}\n{1}\n\n{2}\n{3}\n\n{4}\n{5}\n\n{6}\n{7}\n\n{8}\n{9}\n\n{10}\n{11}\n\n{12}\n{13}\n\n{14}\n{15}'
                .format(d1,i1,d2,i2,d3,i3,d4,i4,d5,i5,d6,i6,d7,i7,d8,i8), reply_markup=MainButtons)
                await state.finish()
    elif message.text == 'Назад':
        await message.answer(text='Меню', reply_markup=MainButtons)
        await state.finish()
    else:
        await message.answer(text='Повторите')

if __name__ == '__main__':
    executor.start_polling(dp, on_startup=on_startup, skip_updates=False)