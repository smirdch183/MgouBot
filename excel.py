import pandas
import openpyxl
import xlrd



# excel_data_df = pandas.read_excel(
#         'test.xlsx',
#         sheet_name='Лист1').fillna(method='ffill', axis=1)
# print(excel_data_df)


# for crange in sheet_ranges.merged_cells:
#     print(crange)
#         rlo, rhi, clo, chi = crange
#         if rowx in xrange(rlo, rhi):
#             if colx in xrange(clo, chi):
#                 heet_ranges.cell_value(rlo,clo)
# print(bokv)
# df = pandas.read_excel('test.xlsx',sheet_name='Лист1')
# print(df)
# for i in bokv:
#     i = str(i)
#     # print(i)
#     # print(sheet_ranges.cell(range=i).value)
#     for row in sheet_ranges[i]:
#         # print(row)
#         for cell in row:
#             # print(cell)
#             cell.value = 'A' # найти начало диапозона
#     sheet_ranges.merge_cells(str(i))


















# from openpyxl import load_workbook
# wb = load_workbook(filename='2_4_курс_45_03_02_Иностранные_языки_и_культуры_стран_изучаемых_языков.xlsx')
# sheet_ranges = wb['Лист1']
# bokv = sheet_ranges.merged_cells.ranges
# from openpyxl.utils.cell import range_boundaries
# for st_name in wb.sheetnames:
#     st = wb[st_name]
#     mcr_coord_list = [mcr.coord for mcr in st.merged_cells.ranges]
    
#     for mcr in mcr_coord_list:
#         min_col, min_row, max_col, max_row = range_boundaries(mcr)
#         top_left_cell_value = st.cell(row=min_row, column=min_col).value
#         st.unmerge_cells(mcr)
#         for row in st.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
#             for cell in row:
#                 cell.value = top_left_cell_value

# wb.save('2_4_курс_45_03_02_Иностранные_языки_и_культуры_стран_изучаемых_языковTest.xlsx')






















# wb.save('test1.xlsm')

# print(wb)


# df = pandas.read_excel('test.xlsx', sheet_name='Лист1')
# excel = openpyxl.load_workbook(filename='test.xlsx')
# sheet = excel.worksheets[0]

# for r in sheet.merged_cells.ranges:
#     cl, rl, cr, rr = r.bounds  # границы объединенной области
#     rl -= 2
#     rr -= 1
#     cl -= 1
#     base_value = df.iloc[rl, cl]
#     df.iloc[rl:rr, cl:cr] = base_value
#     print(base_value)
#     print('-------------')
#     print(df.iloc[rl:rr, cl:cr])
#     print('-------------')
#     print(r)
#     print('-------------')

# print(sheet.merged_cells.ranges)
# print(r.bounds)
# print(df)